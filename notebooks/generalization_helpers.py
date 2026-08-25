from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd


POLICIES = [
    {
        "suffix": "P",
        "code": "pi_P",
        "label": "Capped",
        "title": "Capped policy",
        "cost_col": "poisson_std_10_L6_c1_2",
    },
    {
        "suffix": "E",
        "code": "pi_E",
        "label": "Discounted-pipeline",
        "title": "Discounted-pipeline policy",
        "cost_col": "exponential_std100_L6_c1_2",
    },
    {
        "suffix": "N",
        "code": "pi_N",
        "label": "Hybrid",
        "title": "Hybrid policy",
        "cost_col": "normal_std30_L6_c1_2",
    },
    {
        "suffix": "BS",
        "code": "pi_BS",
        "label": "Base-stock",
        "title": "Base-stock policy",
        "cost_col": "avg_cost_basestock",
    },
]

LEARNED_POLICIES = [policy for policy in POLICIES if policy["suffix"] != "BS"]
INSTANCE_COLUMNS = ["distribution", "demand_mean", "demand_std", "lead_time", "cost_ratio"]

DIST_GROUP_MAP = {
    "beta": "Bounded continuous",
    "cunif": "Bounded continuous",
    "tri": "Bounded continuous",
    "binom": "Bounded discrete",
    "dunif": "Bounded discrete",
    "gamma": "Light/moderate unbounded",
    "geom": "Light/moderate unbounded",
    "negbin": "Light/moderate unbounded",
    "normal": "Light/moderate unbounded",
    "weib": "Light/moderate unbounded",
    "logn": "Heavy-tailed / skewed",
    "pareto": "Heavy-tailed / skewed",
    "zinb": "Heavy-tailed / skewed",
}

DIST_GROUP_ORDER = [
    "Bounded continuous",
    "Bounded discrete",
    "Light/moderate unbounded",
    "Heavy-tailed / skewed",
]

DIST_GROUP_LABELS = [
    "Bounded continuous",
    "Bounded discrete",
    "Light- to moderate-tailed ",
    "Heavy-tailed or skewed",
]


def ensure_generalization_workbook(source_path, output_path):
    """Create the analysis workbook from the raw results workbook."""
    source_path = Path(source_path)
    output_path = Path(output_path)

    from openpyxl import load_workbook

    wb_in = load_workbook(source_path)
    wb_out = load_workbook(source_path)

    keep_sheets = ["S5 merged", "S5 generalization", "S5 train graph", "S5 test graph"]

    for name in list(wb_out.sheetnames):
        if name not in keep_sheets:
            del wb_out[name]

    for name in keep_sheets:
        if name not in wb_out.sheetnames and name in wb_in.sheetnames:
            ws_src = wb_in[name]
            ws_dst = wb_out.create_sheet(title=name)
            for row in ws_src.iter_rows():
                for cell in row:
                    ws_dst[cell.coordinate].value = cell.value
            for merged_range in ws_src.merged_cells.ranges:
                ws_dst.merge_cells(str(merged_range))

    wb_out.save(output_path)
    return output_path


def summarize(frame_or_group_cols, group_cols=None):
    if group_cols is None:
        group_cols = frame_or_group_cols
        frame = None

        # Backward-compatible notebook usage: summarize(["lead_time"]) should
        # work when the caller has a top-level `df` variable in scope.
        import inspect

        caller_frame = inspect.currentframe().f_back
        if caller_frame is not None:
            frame = caller_frame.f_globals.get("df")
    else:
        frame = frame_or_group_cols

    if frame is None:
        raise TypeError(
            "summarize() missing required frame argument; pass summarize(df, [...]) "
            "or define a top-level `df` before calling summarize([...])"
        )

    grouped = frame.groupby(group_cols, dropna=False)
    agg = {f"{policy['suffix']}_CIP": (f"CIP_{policy['suffix']}", "mean") for policy in POLICIES}
    agg.update({f"{policy['suffix']}_AR": (f"AR_{policy['suffix']}", "mean") for policy in POLICIES})
    agg["n"] = ("distribution", "size")
    return grouped.agg(**agg).reset_index().round(2)


def build_overall_table(frame):
    rows = []
    for policy in POLICIES:
        suffix = policy["suffix"]
        rows.append(
            {
                "Policy": policy["code"],
                "Mean CIP (\\%)": frame[f"CIP_{suffix}"].mean(),
                "Median CIP (\\%)": frame[f"CIP_{suffix}"].median(),
                "AR": frame[f"AR_{suffix}"].mean(),
            }
        )
    return pd.DataFrame(rows).round(2)


def build_policy_performance_table(frame, group_col, group_label):
    agg = {
        "Capped_CIP": ("CIP_P", "mean"),
        "Capped_AR": ("AR_P", "mean"),
        "DiscountedPipeline_CIP": ("CIP_E", "mean"),
        "DiscountedPipeline_AR": ("AR_E", "mean"),
        "Hybrid_CIP": ("CIP_N", "mean"),
        "Hybrid_AR": ("AR_N", "mean"),
        "BaseStock_AR": ("AR_BS", "mean"),
        "n": ("distribution", "size"),
    }
    return (
        frame.groupby(group_col, dropna=False)
        .agg(**agg)
        .reset_index()
        .rename(columns={group_col: group_label})
        .round(2)
    )


def assign_quantile_groups(frame, source_col, quantiles, labels, output_col):
    quantile_edges = frame[source_col].quantile(quantiles).to_list()
    bins = [-np.inf, *quantile_edges, np.inf]
    frame[output_col] = pd.cut(frame[source_col], bins=bins, labels=labels, ordered=True)
    return bins


def build_group_panel(frame, group_col, group_values, value_col, labels=None):
    labels = labels or group_values
    return pd.DataFrame(
        {
            label: frame.loc[frame[group_col] == group_value, value_col].reset_index(drop=True)
            for group_value, label in zip(group_values, labels)
        }
    )


def make_legend_handles(median_lw=2.5, mean_size=7, outlier_size=8, median_label="Median"):
    from matplotlib.lines import Line2D

    return [
        Line2D([0], [0], color="black", lw=median_lw, label=median_label),
        Line2D(
            [0],
            [0],
            marker="v",
            color="black",
            markerfacecolor="black",
            markersize=mean_size,
            linestyle="None",
            label="Mean",
        ),
        # Line2D(
        #     [0],
        #     [0],
        #     marker="+",
        #     color="black",
        #     markersize=outlier_size,
        #     linestyle="None",
        #     label="Outlier",
        # ),
    ]


def format_axis(ax, tick_size=12, label_size=12, title_size=14, rotation=0):
    ax.tick_params(axis="x", labelrotation=rotation, labelsize=tick_size)
    ax.tick_params(axis="y", labelsize=tick_size)
    ax.xaxis.label.set_size(label_size)
    ax.yaxis.label.set_size(label_size)
    ax.title.set_fontsize(title_size)


def to_latex(df_table, index=False, caption=None, label=None):
    return df_table.to_latex(index=index, escape=False, caption=caption, label=label)


def boxplot_on_axis(
    ax,
    data,
    cols,
    box_colors,
    labels=None,
    title=None,
    xlabel=None,
    ylabel=None,
    refline=None,
    whis=(5, 95),
    showmeans=True,
    showfliers=True,
    vert=False,
    box_width=0.65,
):
    values = [data[col].dropna() for col in cols]

    bp = ax.boxplot(
        values,
        patch_artist=True,
        showfliers=showfliers,
        showmeans=showmeans,
        widths=[box_width] * len(cols),
        whis=whis,
        boxprops=dict(facecolor="white", lw=1.25, alpha=0.5),
        capprops=dict(lw=2.0, alpha=1, zorder=10),
        whiskerprops=dict(alpha=1, lw=1.5),
        flierprops=dict(alpha=1, lw=0.5, marker="+"),
        medianprops=dict(lw=2.5, color="black"),
        meanprops=dict(alpha=1, marker="v"),
        notch=False,
        vert=vert,
    )

    def safe_set(artists, idx, setter):
        if idx < len(artists):
            setter(artists[idx])

    for i, color in enumerate(box_colors):
        safe_set(bp["boxes"], i, lambda art: (art.set_edgecolor(color), art.set_facecolor(color)))
        safe_set(bp["whiskers"], 2 * i, lambda art: art.set_color(color))
        safe_set(bp["whiskers"], 2 * i + 1, lambda art: art.set_color(color))
        safe_set(bp["caps"], 2 * i, lambda art: art.set_color(color))
        safe_set(bp["caps"], 2 * i + 1, lambda art: art.set_color(color))
        safe_set(bp["medians"], i, lambda art: art.set_color("black"))

        if i < len(bp.get("fliers", [])):
            flier = bp["fliers"][i]
            flier.set_markeredgecolor(color)
            flier.set_markerfacecolor(color)
            flier.set_color(color)
            flier.set_alpha(1)

        if "means" in bp and i < len(bp["means"]):
            mean_marker = bp["means"][i]
            mean_marker.set_markeredgecolor("black")
            mean_marker.set_markerfacecolor("black")
            mean_marker.set_color("black")
            mean_marker.set_alpha(1)

    labels = labels or cols
    tick_positions = range(1, len(labels) + 1)

    if vert:
        ax.set_xticks(tick_positions)
        ax.set_xticklabels(labels)
        if ylabel is not None:
            ax.set_ylabel(ylabel)
        if xlabel is not None:
            ax.set_xlabel(xlabel)
        if refline is not None:
            ax.axhline(refline, linestyle="--", lw=1.2, color="black", alpha=0.8)
    else:
        ax.set_yticks(tick_positions)
        ax.set_yticklabels(labels)
        if xlabel is not None:
            ax.set_xlabel(xlabel)
        if ylabel is not None:
            ax.set_ylabel(ylabel)
        if refline is not None:
            ax.axvline(refline, linestyle="--", lw=1.2, color="black", alpha=0.8)

    if title is not None:
        ax.set_title(title)

    ax.grid(True, linestyle="-", alpha=0.4)
    return bp
